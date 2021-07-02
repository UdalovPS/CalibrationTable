import openpyxl
from insert_table import CalibrationData, RgsTank

class OpenXl():
    def __init__(self, file_name):
        self.file_name = file_name
        # self.open()

    def open_rgs(self):
        try:
            book = openpyxl.open(self.file_name, read_only=True)
            sheets = book.worksheets
            for sheet in sheets:
                if sheet == None: break
                data = self.make_rgs_list(sheet)
            return data
        finally:
            book.close()

    def open_rvs(self):
        try:
            book = openpyxl.open(self.file_name, read_only=True)
            sheets = book.worksheets
            data_list = []
            n = 0
            for sheet in sheets:
                if n == 0:
                    data_dead_belt = self.make_rgs_list(sheet, 1)
                else:
                    data = self.make_rvs_lists(sheet)
                    if data:
                        data_list.append(data)
                n += 1
            # book.close()
            return self.main_data_list(data_list, data_dead_belt)
        finally:
            book.close()

    def make_rgs_list(self, sheet, for_rvs=None):
        level = []
        step = []
        volume = []
        n = 0
        for col in range(0, sheet.max_column):
            for row in range(2, sheet.max_row+1):
                value = sheet[row][col].value
                if value == None: break
                if n == 0:
                    level.append(int(value))
                if n == 1:
                    volume.append(float(value))
                if n == 2:
                    step.append(float(value))
            n += 1
        data = RgsTank().make_rgs_list(level[0], level[1],
                                       volume, step, for_rvs)
        return data

    def make_rvs_lists(self, sheet):
        level = []
        step = []
        volume = []
        n = 0
        for col in range(0, sheet.max_column):
            for row in range(2, sheet.max_row+1):
                value = sheet[row][col].value
                if value == None: break
                if n == 0:
                    level.append(int(value))
                if n == 1:
                    step.append(float(value))
                if n == 2:
                    volume.append(float(value))
            n += 1
        if level and step and volume:
            self.medium_vol(step)
            data = CalibrationData(level[0], level[1],
                                   volume, self.medium_vol(step))
            return (data.main_vlv_vol, data.medium_lvl_vol)

    def medium_vol(self, step):
        step_dict = {}
        for i in range(1, 10):
            step_dict[str(i)] = step[i-1]
        return step_dict

    def main_data_list(self, data_list, data_dead_list):
        list = []
        for super_list in data_list:
            for sub_list in super_list:
                for item in sub_list:
                    data_dead_list.append(item)
        return data_dead_list

if __name__ == '__main__':
    opn = OpenXl("rvs6.xlsx")
    opn.open_rvs()
    
